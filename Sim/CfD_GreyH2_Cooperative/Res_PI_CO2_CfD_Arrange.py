# -*- coding: utf-8 -*-
"""
Created on Wed Nov 19 19:37:15 2025
Corrected version: Ensures ordering of CO2/CfD pairs and fills missing combinations.
@author: Mhdella
"""

import os
import pandas as pd

# === INPUT CONFIGURATION =====================================================


# PI_CO2_CfD_folder = 'CfD_GreyH2_Sub_H2green_Cooperative'
# PI_CO2_CfD_folder = 'CfD_BlueH2_Sub_H2green_Cooperative'
# 
PI_CO2_CfD_folder = "CfD_GreyH2_Cooperative"
# PI_CO2_CfD_folder = "CfD_BlueH2_Cooperative"


# PI_CO2_CfD_folder = "CfD_GreyH2_Hi_LwH2_Cooperative"
# # PI_CO2_CfD_folder = "CfD_BlueH2_Hi_LwH2_Cooperative"


# PI_CO2_CfD_folder = "CfD_GreyH2_Sub_H2green_Cooperative"
# PI_CO2_CfD_folder = "CfD_BlueH2_Sub_H2green_Cooperative"


# PI_CO2_CfD_folder = "CfD_BlueH2_Exg_Competitive"
# PI_CO2_CfD_folder = "CfD_GreyH2_Exg_Competitive"


# # === POLICY COMBINATIONS (Auto-Generated Desired Order) ======================


### if C_EZ_cfd_sw =='Sub_H2green':
# CO2_penalties = [0, 100, 200, 300, 400, 500, 600,    0,   0,   0,   0,   100,  200,  200, 250, 250, 300,  300, 400,  400,  500]
# cfd_values    = [0,   0,   0,   0,   0,   0,   0,   10,  20,  30,  35,    10,   20,   30,  25,  30,  20,   25,  20,   25,   20]


from itertools import product

CO2_penalties_list = [0, 100, 200, 300, 400, 500]
# CO2_penalties_list = [0, 100, 250, 500]

cfd_values_list = [0, 10, 25, 50, 100, 200, 250]
# cfd_values_list = [0, 25, 50, 100, 200, 250]
# cfd_values_list = [0, 50,  100, 150, 200, 250]


# ### if C_EZ_cfd_sw =='SubH2green':
    
# CO2_penalties_list = [0, 100, 250, 400]
# cfd_values_list = [0, 10, 20, 30, 40]




CO2_penalties = []
cfd_values = []

for co2, cfd in product(CO2_penalties_list, cfd_values_list):
    CO2_penalties.append(co2)
    cfd_values .append(cfd)
    
desired_order = list(zip(CO2_penalties, cfd_values))


# st=stop


# Output file names
Emiss_Outs = "Res_CO2_Emissions.xlsx"
Opex_Outs  = "Res_Operational_Cost.xlsx"
OpexSupE_Outs = "Res_Opex_Support_Electricity.xlsx"
OpexSupH_Outs = "Res_Opex_Support_Hydrogen.xlsx"
CfDSupE_Outs  = "Res_CfD_Support_Electricity.xlsx"
CfDSupH_Outs  = "Res_CfD_Support_Hydrogen.xlsx"
H2MC_Outs = "Res_Hydrogen_Marginal_Cost.xlsx"
H2SC_Outs = "Res_Hydrogen_Subsidised_Cost.xlsx"
NPV_Outs  = "Res_NPV.xlsx"
P2GG2G_Outs = "Res_P2G_G2G.xlsx"
H2SubPerMWh_Outs        = "Res_H2_Subsidy_Per_MWh.xlsx"
H2SubTotal_Outs         = "Res_H2_Subsidy_Total.xlsx"
GreenH2SubPerMWh_Outs   = "Res_Green_H2_Subsidy_Per_MWh.xlsx"
GreenH2SubTotal_Outs    = "Res_Green_H2_Subsidy_Total.xlsx"
NPV_wo_CfD_Outs = "Res_NPV_without_CfD_H2.xlsx"
Total_CfD_cost_Outs = "Res_Total_CfD_cost.xlsx"
MAC_Outs = "Res_Marginal_Abatement_Cost.xlsx"



# Base folder
base_path = os.path.join(".", "Output", PI_CO2_CfD_folder, "phi_Variable")

# === NEW: SUMMARY RESULTS FOLDER ============================================

summary_folder = os.path.join(".", "Output", PI_CO2_CfD_folder, "Summary_Results")
os.makedirs(summary_folder, exist_ok=True)

# Full paths for all summary outputs
Emiss_summary    = os.path.join(summary_folder, Emiss_Outs)
Opex_summary     = os.path.join(summary_folder, Opex_Outs)
OpexSupE_summary = os.path.join(summary_folder, OpexSupE_Outs)
OpexSupH_summary = os.path.join(summary_folder, OpexSupH_Outs)
CfDSupE_summary  = os.path.join(summary_folder, CfDSupE_Outs)
CfDSupH_summary  = os.path.join(summary_folder, CfDSupH_Outs)
H2MC_summary = os.path.join(summary_folder, H2MC_Outs)
H2SC_summary = os.path.join(summary_folder, H2SC_Outs)
NPV_summary  = os.path.join(summary_folder, NPV_Outs)
P2GG2G_summary = os.path.join(summary_folder, P2GG2G_Outs)
H2SubPerMWh_summary      = os.path.join(summary_folder, H2SubPerMWh_Outs)
H2SubTotal_summary       = os.path.join(summary_folder, H2SubTotal_Outs)
GreenH2SubPerMWh_summary = os.path.join(summary_folder, GreenH2SubPerMWh_Outs)
GreenH2SubTotal_summary  = os.path.join(summary_folder, GreenH2SubTotal_Outs)
NPV_wo_CfD_summary = os.path.join(summary_folder, NPV_wo_CfD_Outs)
Total_CfD_cost_summary = os.path.join(summary_folder, Total_CfD_cost_Outs)
MAC_summary = os.path.join(summary_folder, MAC_Outs)



price_levels = ["price_high", "price_low"]

# Rows to extract
target_row_emiss = "Total CO2 Emissions [Tonnes]"
target_row_opex  = "Total Operational Cost [m£]"

target_row_opex_sup_e = "Opex-based Support for Electricity [m£]"
target_row_opex_sup_h = "Opex-based Support for Hydrogen [m£]"
target_row_cfd_sup_e  = "CfD-based Support for Electricity [m£]"
target_row_cfd_sup_h  = "CfD-based Support for Hydrogen [m£]"

target_row_h2_mc = "Hydrogen Marginal Cost [£/MWh]"
target_row_h2_sc = "Hydrogen Subsidised Cost [£/MWh]"
target_row_npv   = "Net Present Value (NPV) [b£]"
target_row_p2g   = "P2G [GWh]"
target_row_g2g   = "G2G [GWh]"

target_row_H2_sub_per_MWh = "H2_subsidy_per_MWh (£/MWh)"
target_row_Total_H2_sub   = "Total_H2_subsidy (m£)"
target_row_Green_H2_sub_per_MWh = "Green_H2_subsidy_per_MWh (£/MWh)"
target_row_Total_Green_H2_sub = "Total_Green_H2_subsidy (m£)"
target_row_npv_wo_cfd = "Net Present Value without CfD [b£]"
target_row_total_cfd_cost = "Total CfD cost [b£]"

scenario_names = [
    # "Uniform +25GW/Tech.",
    "High RES–High H2",
    "High RES–Low H2",
    # "High H2–Low BESS"
]

# === STORAGE STRUCTURES ======================================================

summary_emiss_high = []
summary_emiss_low  = []

summary_opex_high = []
summary_opex_low  = []

summary_opex_sup_e_high = []
summary_opex_sup_e_low  = []
summary_opex_sup_h_high = []
summary_opex_sup_h_low  = []

summary_cfd_sup_e_high = []
summary_cfd_sup_e_low  = []
summary_cfd_sup_h_high = []
summary_cfd_sup_h_low  = []

summary_h2_mc_high = []
summary_h2_mc_low  = []

summary_h2_sc_high = []
summary_h2_sc_low  = []

summary_npv_high = []
summary_npv_low  = []

summary_p2g_g2g_high = []
summary_p2g_g2g_low  = []

summary_H2_sub_per_MWh_high = []
summary_H2_sub_per_MWh_low  = []

summary_Total_H2_sub_high = []
summary_Total_H2_sub_low  = []

summary_Green_H2_sub_per_MWh_high = []
summary_Green_H2_sub_per_MWh_low  = []

summary_Total_Green_H2_sub_high = []
summary_Total_Green_H2_sub_low  = []

summary_npv_wo_cfd_high = []
summary_npv_wo_cfd_low  = []

summary_total_cfd_cost_high = []
summary_total_cfd_cost_low  = []
# === MAIN LOOP ===============================================================

if not os.path.exists(base_path):
    raise FileNotFoundError(f"Base folder not found: {base_path}")

for folder_name in os.listdir(base_path):
    folder_path = os.path.join(base_path, folder_name)
    if not os.path.isdir(folder_path):
        continue

    # ==============================================================
    # FIXED: Correct parsing of folder names
    # Expected format: "CO2_<value>_CfD_<value>"
    # ==============================================================

    # PI_CfD_<CfD_value>_CO2_<CO2_value>
    parts = folder_name.split("_")
    
    try:
        cfd_value   = int(parts[2])   # after "CfD"
        CO2_penalty = int(parts[4])   # after "CO2"
    except (IndexError, ValueError):
        print(f"⚠ Skipping folder with unexpected name: {folder_name}")
        continue


    for price_level in price_levels:

        excel_path = os.path.join(folder_path, price_level, "All_simulations_results.xlsx")
        if not os.path.exists(excel_path):
            print(f"⚠ File not found: {excel_path}")
            continue

        df = pd.read_excel(excel_path, sheet_name="Simulation Summary", index_col=0)
        df_policy = pd.read_excel(excel_path, sheet_name="Policy Support Summary", index_col=0)

        # --- Emissions ---
        if target_row_emiss in df.index:
            values = df.loc[target_row_emiss, scenario_names].tolist()
            entry = [CO2_penalty, cfd_value] + values
            (summary_emiss_high if price_level=="price_high" else summary_emiss_low).append(entry)

        # --- Operational Cost ---
        if target_row_opex in df.index:
            values = df.loc[target_row_opex, scenario_names].tolist()
            entry = [CO2_penalty, cfd_value] + values
            (summary_opex_high if price_level=="price_high" else summary_opex_low).append(entry)

        # --- Hydrogen Marginal Cost ---
        if target_row_h2_mc in df.index:
            values = df.loc[target_row_h2_mc, scenario_names].tolist()
            entry = [CO2_penalty, cfd_value] + values
            (summary_h2_mc_high if price_level=="price_high" else summary_h2_mc_low).append(entry)
        
        # --- Hydrogen Subsidised Cost ---
        if target_row_h2_sc in df.index:
            values = df.loc[target_row_h2_sc, scenario_names].tolist()
            entry = [CO2_penalty, cfd_value] + values
            (summary_h2_sc_high if price_level=="price_high" else summary_h2_sc_low).append(entry)
        
        
        # --- Net Present Value ---
        if target_row_npv in df.index:
            values = df.loc[target_row_npv, scenario_names].tolist()
            entry = [CO2_penalty, cfd_value] + values
            (summary_npv_high if price_level=="price_high" else summary_npv_low).append(entry)
        
        # --- Net Present Value without CfD ---
        if target_row_npv_wo_cfd in df.index:
            values = df.loc[target_row_npv_wo_cfd, scenario_names].tolist()
            entry = [CO2_penalty, cfd_value] + values
            (summary_npv_wo_cfd_high if price_level=="price_high" else summary_npv_wo_cfd_low).append(entry)

        # --- Total CfD cost ---
        if target_row_total_cfd_cost in df.index:
            values = df.loc[target_row_total_cfd_cost, scenario_names].tolist()
            entry = [CO2_penalty, cfd_value] + values
            (summary_total_cfd_cost_high if price_level=="price_high" else summary_total_cfd_cost_low).append(entry)


        # --- P2G + G2G ---
        if (target_row_p2g in df.index) and (target_row_g2g in df.index):
            combined_values = []
            
            for sc in scenario_names:
                combined_values.append(df.loc[target_row_p2g, sc])
                combined_values.append(df.loc[target_row_g2g, sc])
            entry = [CO2_penalty, cfd_value] + combined_values
            (summary_p2g_g2g_high if price_level=="price_high" else summary_p2g_g2g_low).append(entry)

        # --- Policy Support ---
        policy_map = {
            target_row_opex_sup_e: (summary_opex_sup_e_high, summary_opex_sup_e_low),
            target_row_opex_sup_h: (summary_opex_sup_h_high, summary_opex_sup_h_low),
            target_row_cfd_sup_e:  (summary_cfd_sup_e_high,  summary_cfd_sup_e_low),
            target_row_cfd_sup_h:  (summary_cfd_sup_h_high,  summary_cfd_sup_h_low),
        }

        for row_name, (high_list, low_list) in policy_map.items():
            if row_name in df_policy.index:
                values = df_policy.loc[row_name, scenario_names].tolist()
                entry = [CO2_penalty, cfd_value] + values
                (high_list if price_level=="price_high" else low_list).append(entry)
            else:
                print(f"⚠ Policy row '{row_name}' missing in: {excel_path}")
                
                
        # --- Additional Hydrogen Subsidy Rows ---
        hydro_policy_map = {
            target_row_H2_sub_per_MWh: (summary_H2_sub_per_MWh_high, summary_H2_sub_per_MWh_low),
            target_row_Total_H2_sub:   (summary_Total_H2_sub_high,   summary_Total_H2_sub_low),
            target_row_Green_H2_sub_per_MWh: (summary_Green_H2_sub_per_MWh_high, summary_Green_H2_sub_per_MWh_low),
            target_row_Total_Green_H2_sub:   (summary_Total_Green_H2_sub_high,   summary_Total_Green_H2_sub_low),
        }
        
        for row_name, (high_list, low_list) in hydro_policy_map.items():
            if row_name in df_policy.index:
                values = df_policy.loc[row_name, scenario_names].tolist()
                entry = [CO2_penalty, cfd_value] + values
                (high_list if price_level=="price_high" else low_list).append(entry)
            else:
                print(f"⚠ Policy row '{row_name}' missing in: {excel_path}")
                

                
        

# === DESIRED ORDER ===========================================================


order_index = pd.MultiIndex.from_tuples(desired_order, names=["CO2 Price","CfD Price"])

columns = ["CO2 Price", "CfD Price"] + scenario_names

# === SAVE ORDERED EXCEL FILES ===============================================
dfs = {
    Emiss_summary: (summary_emiss_high, summary_emiss_low),
    Opex_summary: (summary_opex_high, summary_opex_low),
    OpexSupE_summary: (summary_opex_sup_e_high, summary_opex_sup_e_low),
    OpexSupH_summary: (summary_opex_sup_h_high, summary_opex_sup_h_low),
    CfDSupE_summary: (summary_cfd_sup_e_high, summary_cfd_sup_e_low),
    CfDSupH_summary: (summary_cfd_sup_h_high, summary_cfd_sup_h_low),
    H2MC_summary: (summary_h2_mc_high, summary_h2_mc_low),
    H2SC_summary: (summary_h2_sc_high, summary_h2_sc_low),
    NPV_summary: (summary_npv_high, summary_npv_low),
    NPV_wo_CfD_summary: (summary_npv_wo_cfd_high, summary_npv_wo_cfd_low),
    Total_CfD_cost_summary: (summary_total_cfd_cost_high, summary_total_cfd_cost_low),
    P2GG2G_summary: (summary_p2g_g2g_high, summary_p2g_g2g_low),
    

    H2SubPerMWh_summary:      (summary_H2_sub_per_MWh_high, summary_H2_sub_per_MWh_low),
    H2SubTotal_summary:       (summary_Total_H2_sub_high,   summary_Total_H2_sub_low),
    GreenH2SubPerMWh_summary: (summary_Green_H2_sub_per_MWh_high, summary_Green_H2_sub_per_MWh_low),
    GreenH2SubTotal_summary:  (summary_Total_Green_H2_sub_high,   summary_Total_Green_H2_sub_low),
}



from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# Define a list of fills for alternating/fixed scenarios
scenario_fills = [
    PatternFill(start_color="FFEEEE", end_color="FFEEEE", fill_type="solid"),  # scenario 1
    PatternFill(start_color="EEEEFF", end_color="EEEEFF", fill_type="solid"),  # scenario 2
    PatternFill(start_color="EEFFEE", end_color="EEFFEE", fill_type="solid"),  # scenario 3
    PatternFill(start_color="FFFFEE", end_color="FFFFEE", fill_type="solid"),  # scenario 4
]

for file_path, (high_list, low_list) in dfs.items():

    # === Set correct columns BEFORE building dataframes ===
    if file_path == P2GG2G_summary:
        # Two columns per scenario: P2G & G2G
        scenario_cols = []
        for sc in scenario_names:
            scenario_cols += [f"{sc} - P2G", f"{sc} - G2G"]
        columns = ["CO2 Price", "CfD Price"] + scenario_cols
    else:
        # Normal files: one value per scenario
        columns = ["CO2 Price", "CfD Price"] + scenario_names

    # === Build DataFrames ===
    df_high = pd.DataFrame(high_list, columns=columns)
    df_low  = pd.DataFrame(low_list, columns=columns)

    # === Reorder rows ===
    desired_df = pd.DataFrame(desired_order, columns=["CO2 Price","CfD Price"])
    df_high = desired_df.merge(df_high, on=["CO2 Price","CfD Price"], how="left").fillna(0)
    df_low  = desired_df.merge(df_low,  on=["CO2 Price","CfD Price"], how="left").fillna(0)

    # === Save initial Excel ===
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df_high.to_excel(writer, sheet_name="High_Price", index=False)
        df_low.to_excel(writer, sheet_name="Low_Price", index=False)

    # === Apply styling ONLY for P2G/G2G file ===
    if file_path == P2GG2G_summary:
        wb = load_workbook(file_path)
        for sheet_name in ["High_Price", "Low_Price"]:
            ws = wb[sheet_name]

            # ---- Apply fill per scenario ----
            for i, sc in enumerate(scenario_names):
                fill = scenario_fills[i % len(scenario_fills)]
                p2g_col = 3 + i*2  # P2G column index (1-based)
                g2g_col = 4 + i*2  # G2G column index

                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=p2g_col).fill = fill
                    ws.cell(row=row, column=g2g_col).fill = fill

                    # Font colors
                    ws.cell(row=row, column=p2g_col).font = Font(color="008000")  # green P2G
                    ws.cell(row=row, column=g2g_col).font = Font(color="0000FF")  # blue G2G

            # ---- Optional: Align headers center ----
            for col in range(1, ws.max_column + 1):
                ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")

        # === Save workbook with styles ===
        wb.save(file_path)



# =============================================================================
# === MARGINAL ABATEMENT COST (MAC) CALCULATION ================================
# =============================================================================

def compute_mac(cost_df, emiss_df, scenario_names, eps=1e-3):
    """
    Compute Marginal Abatement Cost (MAC) for each policy scenario.

    MAC_i = (C_i - C_baseline) / (E_baseline - E_i) [£/tCO2]

    Baseline: CO2 = 0, CfD = 0

    Parameters:
        cost_df      : Operational cost dataframe (m£)
        emiss_df     : CO2 emissions dataframe (tCO2)
        scenario_names : list of scenario names (columns)
        eps          : minimal ΔE threshold to avoid numerical spikes

    Returns:
        mac_df : DataFrame with MAC values (£/tCO2)
    """

    mac_df = cost_df.copy()

    # --- Identify baseline row ---
    baseline_mask = (cost_df["CO2 Price"] == 0) & (cost_df["CfD Price"] == 0)
    if baseline_mask.sum() == 0:
        raise ValueError("Baseline (CO2=0, CfD=0) not found for MAC calculation.")

    for sc in scenario_names:
        # Baseline values
        C_base = cost_df.loc[baseline_mask, sc].values[0] * 1e6  # m£ → £
        E_base = emiss_df.loc[baseline_mask, sc].values[0]

        mac_values = []

        for _, row in cost_df.iterrows():
            # Policy case
            C_i = row[sc] * 1e6
            E_i = emiss_df.loc[
                (emiss_df["CO2 Price"] == row["CO2 Price"]) &
                (emiss_df["CfD Price"] == row["CfD Price"]),
                sc
            ].values[0]

            delta_E = E_base - E_i
            delta_C = C_i - C_base

            # Avoid divide-by-zero or very small abatement
            if abs(delta_E) <= eps:
                # mac_values.append(float("nan"))  # undefined MAC
                mac_values.append(0)  # undefined MAC

            else:
                mac_values.append(delta_C / delta_E)

        mac_df[sc] = mac_values

    return mac_df


# === Load existing summary results ===
emiss_xls = pd.ExcelFile(Emiss_summary)
opex_xls  = pd.ExcelFile(Opex_summary)

mac_results = {}

for sheet in ["High_Price", "Low_Price"]:
    emiss_df = emiss_xls.parse(sheet)
    opex_df  = opex_xls.parse(sheet)

    mac_df = compute_mac(
        cost_df=opex_df,
        emiss_df=emiss_df,
        scenario_names=scenario_names
    )

    mac_results[sheet] = mac_df


# === Save MAC results to Excel ===
with pd.ExcelWriter(MAC_summary, engine="openpyxl") as writer:
    mac_results["High_Price"].to_excel(writer, sheet_name="High_Price", index=False)
    mac_results["Low_Price"].to_excel(writer, sheet_name="Low_Price", index=False)

print(f"✅ Marginal Abatement Cost file created: {MAC_summary}")

